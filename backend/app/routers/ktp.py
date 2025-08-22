from fastapi import APIRouter, HTTPException, Form, Request
from fastapi.responses import FileResponse
from typing import List
import os
import time
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from app.core.config import settings

# Основной роутер для API v2
router = APIRouter(prefix="/api/ktp", tags=["КТП генератор"])

# Legacy роутер для совместимости
legacy_router = APIRouter(prefix="/api", tags=["КТП генератор (Legacy)"])

def parse_date(date_str):
    """Парсит дату в формате DD.MM.YYYY"""
    try:
        return datetime.strptime(date_str, '%d.%m.%Y')
    except ValueError:
        return None

def format_date(date):
    """Форматирует дату в DD.MM"""
    return date.strftime('%d.%m')

def is_weekend(date):
    """Проверяет, является ли дата выходным"""
    return date.weekday() >= 5

def generate_schedule(start_date, end_date, weekdays, lessons_per_day, holidays, vacation_dates):
    """Генерирует расписание уроков"""
    schedule = []
    current_date = start_date
    
    while current_date <= end_date:
        # Пропускаем выходные
        if current_date.weekday() in weekdays:
            # Проверяем, не праздник ли это
            date_str = current_date.strftime('%d.%m.%Y')
            if date_str not in holidays and date_str not in vacation_dates:
                # Добавляем дату в расписание столько раз, сколько уроков в этот день
                weekday_index = current_date.weekday()
                if weekday_index < len(lessons_per_day):
                    lessons_count = lessons_per_day[weekday_index]
                    for _ in range(lessons_count):
                        schedule.append({
                            'date': format_date(current_date)
                        })
        
        current_date += timedelta(days=1)
    
    return schedule

def create_excel_schedule(schedule, filename):
    """Создает Excel файл с расписанием"""
    # Создаем DataFrame
    df = pd.DataFrame(schedule)
    
    # Создаем Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Расписание"
    
    # Заголовки
    headers = ['Дата']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Данные
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)
    
    # Автоматическая ширина столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем файл
    filepath = os.path.join(settings.temp_dir, f"{filename}.xlsx")
    wb.save(filepath)
    return filepath

@router.post("/generate")
async def generate_ktp_schedule(
    request: Request,
    start_date: str = Form(...),
    end_date: str = Form(...),
    weekdays: List[int] = Form(...),
    lessons_per_day: List[int] = Form(...),
    holidays: List[str] = Form(default=[]),
    vacation: List[str] = Form(default=[]),
    file_name: str = Form("schedule")
):
    """
    Генерация календарно-тематического планирования
    
    **Параметры:**
    - `start_date`: Начальная дата (YYYY-MM-DD)
    - `end_date`: Конечная дата (YYYY-MM-DD)
    - `weekdays`: Рабочие дни недели (0-6, где 0=понедельник)
    - `lessons_per_day`: Количество уроков в каждый день недели
    - `holidays`: Список праздничных дат (DD.MM.YYYY)
    - `vacation`: Список дат каникул (DD.MM.YYYY)
    - `file_name`: Имя файла для сохранения
    """
    
    start_time = time.time()
    
    try:
        # Парсим даты
        start = datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y-%m-%d')
        
        if start >= end:
            raise HTTPException(
                status_code=400,
                detail="Начальная дата должна быть раньше конечной"
            )
        
        # Валидация рабочих дней
        if not weekdays:
            raise HTTPException(
                status_code=400,
                detail="Необходимо выбрать хотя бы один рабочий день"
            )
        
        # Проверяем, что количество уроков соответствует дням недели
        if len(lessons_per_day) != 7:
            raise HTTPException(
                status_code=400,
                detail="Количество уроков должно быть указано для всех 7 дней недели"
            )
        
        # Генерируем расписание
        schedule = generate_schedule(
            start, end, weekdays, lessons_per_day, holidays, vacation
        )
        
        if not schedule:
            raise HTTPException(
                status_code=400,
                detail="Не удалось сгенерировать расписание. Проверьте параметры."
            )
        
        # Создаем Excel файл
        excel_path = create_excel_schedule(schedule, file_name)
        
        # Измеряем время обработки
        processing_time = int((time.time() - start_time) * 1000)
        
        # Возвращаем файл
        return FileResponse(
            excel_path,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename=f"{file_name}.xlsx",
            headers={
                'X-Processing-Time': str(processing_time),
                'X-Lessons-Count': str(len(schedule))
            }
        )
        
    except ValueError as e:
        raise HTTPException(
            status_code=400,
            detail=f"Ошибка формата даты: {str(e)}"
        )
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка генерации расписания: {str(e)}"
        )

# Legacy endpoint
@legacy_router.post("/ktp-generator")
async def legacy_ktp_generator(
    request: Request,
    start_date: str = Form(...),
    end_date: str = Form(...),
    weekdays: List[int] = Form(...),
    lessons_per_day: List[int] = Form(...),
    holidays: List[str] = Form(default=[]),
    vacation: List[str] = Form(default=[]),
    file_name: str = Form("schedule")
):
    """Legacy endpoint для совместимости"""
    return await generate_ktp_schedule(
        request, start_date, end_date, weekdays, 
        lessons_per_day, holidays, vacation, file_name
    ) 